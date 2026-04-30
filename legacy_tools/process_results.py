import os
import shutil
import openpyxl

result_path = 'result.md'
folder_path = '方案评审'
reviewed_folder = '已审核文件夹'
excel_path = '检查结果整理.xlsx'
backup_excel_path = '检查结果整理_备份.xlsx'

# 1. Parse result.md
reviewed_files = []
missing_projects = []

current_section = 0  # 1: reviewed, 2: unreviewed, 3: missing

with open(result_path, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if line.startswith('## 一、已评审的'):
            current_section = 1
        elif line.startswith('## 二、未评审的'):
            current_section = 2
        elif line.startswith('## 三、有评审意见但缺失'):
            current_section = 3
            
        if line.startswith('- **') and line.endswith('**'):
            # Extraction: "- **filename**" -> "filename"
            name = line[4:-2]
            if current_section == 1:
                reviewed_files.append(name)
            elif current_section == 3:
                missing_projects.append(name)

print(f"Parsed {len(reviewed_files)} reviewed files and {len(missing_projects)} missing projects from result.md.")

# 2. Create directory and Move files
if not os.path.exists(reviewed_folder):
    os.makedirs(reviewed_folder)

moved_count = 0
for file_name in reviewed_files:
    src = os.path.join(folder_path, file_name)
    dst = os.path.join(reviewed_folder, file_name)
    if os.path.exists(src):
        shutil.move(src, dst)
        moved_count += 1
    else:
        print(f"Warning: File {src} not found!")
print(f"Moved {moved_count} files to {reviewed_folder}.")

# 3. Annotate Excel
if not os.path.exists(backup_excel_path):
    shutil.copy(excel_path, backup_excel_path)
    
wb = openpyxl.load_workbook(excel_path)
annotated_count = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Check if '备注' exists in row 1
    remark_col = None
    for col_idx in range(1, ws.max_column + 2):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val == '备注':
            remark_col = col_idx
            break
            
    if remark_col is None:
        remark_col = ws.max_column + 1
        ws.cell(row=1, column=remark_col).value = '备注'
        
    for row_idx in range(2, ws.max_row + 1):
        # We need to see if any cell in this row matches one of missing_projects
        # Or more specifically, if the project name matches
        row_matched = False
        for col_idx in range(1, ws.max_column + 1):
            val = str(ws.cell(row=row_idx, column=col_idx).value or '').strip()
            if val in missing_projects:
                row_matched = True
                break
                
        if row_matched:
            ws.cell(row=row_idx, column=remark_col).value = '有评审意见但缺失对应的评审文件'
            annotated_count += 1

wb.save(excel_path)
print(f"Annotated {annotated_count} projects in Excel.")
