import os
import shutil
import openpyxl
from difflib import SequenceMatcher

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

folder_path = '方案评审'
excel_path = '检查结果整理.xlsx'
backup_excel_path = '检查结果整理_备份.xlsx'
reviewed_folder = '已审核文件夹'

# Create reviewed folder if it doesn't exist
if not os.path.exists(reviewed_folder):
    os.makedirs(reviewed_folder)

# Backup Excel
shutil.copy(excel_path, backup_excel_path)

# Get files
try:
    files = os.listdir(folder_path)
    files = [f for f in files if os.path.isfile(os.path.join(folder_path, f))]
except Exception as e:
    files = []

# Process Excel with openpyxl to preserve formatting
wb = openpyxl.load_workbook(excel_path)

# We have to keep track of matched rows and matched files
matched_files = set()
matched_cells = set() # Store tuples of (sheet_name, row_idx) that are matched

# We iterate sheets to collect data for fuzzy matching
excel_data = [] # List of dict: sheet_name, row_idx, proj_name, opinion, has_opinion
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Find column indices (1-indexed)
    header_row = 1
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            col_map[str(val).strip()] = col_idx
            
    proj_col = col_map.get('方案/白单名称', col_map.get('项目', col_map.get('工程名称', col_map.get('文件名'))))
    opinion_col = col_map.get('意见', col_map.get('具体意见'))
    has_opinion_col = col_map.get('是否存在意见')
    
    # If standard columns not found, use heuristics:
    if proj_col is None:
        for val, idx in col_map.items():
            if '项目' in val or '名称' in val or '方案' in val:
                proj_col = idx
                break
                
    if proj_col is not None:
        for row_idx in range(2, ws.max_row + 1):
            proj_val = str(ws.cell(row=row_idx, column=proj_col).value or '').strip()
            opinion_val = str(ws.cell(row=row_idx, column=opinion_col).value or '').strip() if opinion_col else ''
            has_opinion_val = str(ws.cell(row=row_idx, column=has_opinion_col).value or '').strip() if has_opinion_col else ''
            
            has_opinion = has_opinion_val == '是' or opinion_val != ''
            
            if proj_val:
                excel_data.append({
                    'sheet_name': sheet_name,
                    'row_idx': row_idx,
                    'proj_name': proj_val,
                    'has_opinion': has_opinion
                })

# Fuzzy match
reviewed_files_list = []
for file_name in files:
    name_no_ext = os.path.splitext(file_name)[0]
    best_match_idx = -1
    best_score = 0
    
    for j, data in enumerate(excel_data):
        proj_name = data['proj_name']
        if proj_name in name_no_ext or name_no_ext in proj_name:
            score = 1.0
        else:
            score = similar(name_no_ext, proj_name)
            
        if score > 0.6 and score > best_score:
            best_score = score
            best_match_idx = j
            
    if best_score > 0.6:
        matched_files.add(file_name)
        data = excel_data[best_match_idx]
        matched_cells.add((data['sheet_name'], data['row_idx']))
        
        if data['has_opinion']:
            reviewed_files_list.append(file_name)

# 1. Move reviewed files
moved_count = 0
for file_name in reviewed_files_list:
    src_path = os.path.join(folder_path, file_name)
    dst_path = os.path.join(reviewed_folder, file_name)
    if os.path.exists(src_path):
        shutil.move(src_path, dst_path)
        moved_count += 1
print(f"Moved {moved_count} files to {reviewed_folder}")

# 2. Annotate missing files in Excel
annotated_count = 0
for data in excel_data:
    if (data['sheet_name'], data['row_idx']) not in matched_cells:
        if data['has_opinion']:
            # Project has opinion but no file matched
            ws = wb[data['sheet_name']]
            # Find or create a '备注' column
            header_row = 1
            remark_col = None
            for col_idx in range(1, ws.max_column + 2):
                val = ws.cell(row=header_row, column=col_idx).value
                if val == '备注':
                    remark_col = col_idx
                    break
            
            if remark_col is None:
                remark_col = ws.max_column + 1
                ws.cell(row=header_row, column=remark_col).value = '备注'
                
            # Add remark
            ws.cell(row=data['row_idx'], column=remark_col).value = '有评审意见但缺失评审文件'
            annotated_count += 1

# Save Excel
wb.save(excel_path)
print(f"Annotated {annotated_count} projects in {excel_path}")
