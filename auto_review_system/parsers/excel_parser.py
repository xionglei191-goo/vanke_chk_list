import openpyxl
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def parse_excel_bill(file_path):
    """
    解析 Excel 报价清单，以最宽松的策略提取文本。
    """
    results = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            header_row = 1
            col_map = {}
            for row_idx in range(1, min(15, ws.max_row + 1)):
                for col_idx in range(1, ws.max_column + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, str):
                        # 只要出现项目/金额/数量/特征任何一个，大概率就是表头
                        if any(kw in val for kw in ['特征', '名称', '内容', '单价', '合价', '数量']):
                            header_row = row_idx
                            break
                if header_row != 1:
                    break
                    
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                if val:
                    col_map[str(val).strip()] = col_idx
                    
            # Try combinations of common names
            item_cols = []
            for k, v in col_map.items():
                if any(x in k for x in ['特征', '项目', '名称', '内容', '明细', '描述', '说明', '意见']):
                    item_cols.append(v)
            
            # If no obvious keyword found, default to column 2 or 3 (usually contains text)
            if not item_cols and ws.max_column >= 2:
                item_cols = [2]
                
            qty_col = None
            price_col = None
            for k, v in col_map.items():
                if '数量' in k or '量' in k:
                    qty_col = v
                if '单' in k or '价' in k:
                    price_col = v
            
            items = []
            # 只要有文本列，我们就强行提取
            if item_cols:
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    # Combine all textual item columns into one string
                    item_name_parts = []
                    for c in item_cols:
                        val = str(ws.cell(row=row_idx, column=c).value or '').strip()
                        if val and val != 'None':
                            item_name_parts.append(val)
                            
                    item_name = " | ".join(item_name_parts)
                    if not item_name:
                        continue
                        
                    qty = ws.cell(row=row_idx, column=qty_col).value if qty_col else 1
                    price = ws.cell(row=row_idx, column=price_col).value if price_col else 0
                    
                    try:
                        items.append({
                            'row': row_idx,
                            'name': item_name,
                            'quantity': float(qty) if qty is not None else 1,
                            'price': float(price) if price is not None else 0
                        })
                    except (ValueError, TypeError):
                        items.append({
                            'row': row_idx,
                            'name': item_name,
                            'quantity': 1,
                            'price': 0
                        })
            
            proj_name = str(ws.cell(row=1, column=1).value or sheet_name).strip()
            if len(proj_name) < 2 or '表' in proj_name:
                proj_name = sheet_name

            results.append({
                'sheet_name': sheet_name,
                'project_name': proj_name,
                'items': items
            })
    except Exception as e:
        print(f"Error parsing excel {file_path}: {e}")
        
    return results

def parse_excel_as_scheme_chunks(excel_path):
    """
    将施工方案型 Excel 转换为模型识别的 Chunks。
    把大段文字合并为段落，以便供给方案审查特工 (Agents 1-7)。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return f"Error loading Excel scheme: {e}"
        
    chunks = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_blocks = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # 将该行有效单元格拼接
            row_text = " | ".join([str(v) for v in row if v is not None and str(v).strip()])
            if row_text:
                text_blocks.append(f"[第{row_idx}行]: {row_text}")
        
        # 将整个 Sheet 的文本合并为一个 chunk，如果过大可以截断，此处为了全链路扫描采用合并策略
        if text_blocks:
            chunk_text = "\\n".join(text_blocks)
            chunks.append({
                "heading": f"Excel表格结构拆解 [{sheet_name}]",
                "text": chunk_text
            })
            
    if not chunks:
        chunks.append({"heading": "Excel文件空", "text": "文件中未找到有效文本"})
        
    return chunks
